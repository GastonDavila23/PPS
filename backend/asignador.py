"""
================================================================================
ARCHIVO: asignador.py
================================================================================
PROPÓSITO:
Este archivo es el servidor principal del backend (API REST) para el Sistema
Asignador de Profesores, construido con Flask.

MANEJA:
- Conexión a la base de datos (SQLite).
- Todos los endpoints (rutas) de la API.
- Lógica de cálculo de asignaciones geográficas (recalcular_y_guardar_asignaciones).
- Endpoints para el frontend:
    - /api/asignaciones: Devuelve la lista de asignaciones con filtros y paginación.
    - /api/cargar-planillas: Recibe los archivos Excel, los procesa y recalcula todo.
    - /api/descargar-excel: Genera y devuelve el reporte de Excel con filtros y estilos.
- Endpoints de gestión de usuarios y roles para el panel de administración.
================================================================================
"""

# --- Importaciones de librerías ---
import pandas as pd                 # Para manejar datos en DataFrames (similar a Excel)
import sqlite3                      # Para la base de datos SQLite
import io                           # Para manejar la salida del archivo Excel en memoria
import math                         # Para cálculos matemáticos (ceil para paginación)
from flask import Flask, jsonify, request, send_file  # El framework principal del backend
from flask_cors import CORS         # Para permitir que el frontend (en otro dominio) se conecte
from geopy.distance import geodesic # Para calcular la distancia entre coordenadas (lat, lon)
from procesador_excel import procesar_archivos_y_actualizar # Importa la lógica de limpieza de Excel
import numpy as np                  # Usado por Pandas

# --- Importaciones para Estilos de Excel ---
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# --- Configuración Inicial de la Aplicación Flask ---
app = Flask(__name__) # Inicializa la aplicación Flask
CORS(app)             # Habilita CORS para todas las rutas

# --- Funciones de Base de Datos ---

def get_db_connection():
    """
    Función de utilidad para crear y retornar una conexión a la base de datos.
    Usa `sqlite3.Row` para que se pueda acceder a los resultados por nombre de columna.
    """
    conn = sqlite3.connect('asignador.db')
    conn.row_factory = sqlite3.Row
    return conn

# --- Lógica Principal del Negocio ---

def recalcular_y_guardar_asignaciones():
    """
    EL CEREBRO DEL SISTEMA.
    Esta función se llama después de cargar nuevas planillas.
    1. Lee todos los datos limpios de la tabla 'escuelas_data'.
    2. Calcula la asignación óptima (más cercana) para cada profesor.
    3. Guarda los resultados en la tabla 'resultados_asignacion'.
    """
    try:
        # 1. Leer datos de escuelas desde la BD
        conn = get_db_connection()
        df = pd.read_sql_query("SELECT * FROM escuelas_data", conn)
        if df.empty:
            conn.execute("DELETE FROM resultados_asignacion") # Limpia resultados si no hay datos
            conn.commit()
            print("INFO: No hay datos de escuelas, se han limpiado las asignaciones existentes.")
            return
    except Exception as e:
        print(f"ERROR: No se pudo leer la tabla de escuelas. {e}")
        return
    finally:
        if conn:
            conn.close()

    destinos_disponibles = df.copy() # Copia el df para ir "quitando" destinos ya asignados
    lista_resultados = []           # Lista para guardar los resultados
    
    # --- Definición de Observaciones estándar ---
    OBS_FALTA_DATOS = "No Asignada (Faltan Datos Geo)"
    OBS_SIN_CANDIDATOS = "No Asignada (Sin Candidatos)"
    OBS_MAYOR_30KM = "No Asignada (Candidatos > 30km)"
    OBS_0_5KM = "Asignado (0-5 km)"
    OBS_5_10KM = "Asignado (5-10 km)"
    OBS_10_30KM = "Asignado (10-30 km)"

    # --- 2. Iterar por cada profesor (origen) ---
    for _, profesor_origen in df.iterrows():
        origen_info = profesor_origen.to_dict()
        
        # Si el origen no tiene lat/lon, no se puede asignar
        if pd.isna(origen_info.get('Latitud')) or pd.isna(origen_info.get('Longitud')):
            resultado = {
                "origen": origen_info, "destino": origen_info,
                "Distancia_KM": 0.0, "Observaciones": OBS_FALTA_DATOS
            }
            lista_resultados.append(resultado)
            continue
        
        # --- 3. Buscar destinos (candidatos) ---
        candidatos = []
        for _, posible_destino in destinos_disponibles.iterrows():
            
            # Condición de asignación:
            # - El destino debe tener lat/lon
            # - No debe ser la misma escuela (CUE diferente)
            # - DEBE COINCIDIR EL TURNO
            if (pd.notna(posible_destino.get('Latitud')) and pd.notna(posible_destino.get('Longitud')) and
                origen_info.get('CUE') != posible_destino.get('CUE') and
                origen_info.get('Turno') == posible_destino.get('Turno')):
                
                # Calcular distancia geodésica (la más precisa)
                dist = geodesic(
                    (origen_info['Latitud'], origen_info['Longitud']),
                    (posible_destino['Latitud'], posible_destino['Longitud'])
                ).kilometers
                candidatos.append({'destino_serie': posible_destino, 'distancia': dist, 'index': posible_destino.name})
        
        # Si no hay candidatos que cumplan la condición
        if not candidatos:
            resultado = {
                "origen": origen_info, "destino": origen_info,
                "Distancia_KM": 0.0, "Observaciones": OBS_SIN_CANDIDATOS
            }
            lista_resultados.append(resultado)
            continue

        # --- 4. Encontrar la mejor opción (la más cercana) ---
        mejor_opcion = sorted(candidatos, key=lambda x: x['distancia'])[0]
        distancia_final = mejor_opcion['distancia']
        
        # --- 5. Clasificar la asignación por rango de distancia ---
        if distancia_final < 5:
            observacion = OBS_0_5KM
        elif 5 <= distancia_final < 10:
            observacion = OBS_5_10KM
        elif 10 <= distancia_final <= 30:
            observacion = OBS_10_30KM
        else:
            # Si el más cercano está a más de 30km, no se asigna
            resultado = {
                "origen": origen_info, "destino": origen_info,
                "Distancia_KM": 0.0, "Observaciones": OBS_MAYOR_30KM
            }
            lista_resultados.append(resultado)
            continue
        
        # --- 6. Guardar asignación exitosa ---
        destino_info = mejor_opcion['destino_serie'].to_dict()
        resultado = {
            "origen": origen_info, "destino": destino_info,
            "Distancia_KM": distancia_final, "Observaciones": observacion
        }
        # Elimina el destino de la lista de disponibles para que no se asigne 2 veces
        destinos_disponibles.drop(mejor_opcion['index'], inplace=True, errors='ignore')
        lista_resultados.append(resultado)

    # --- 7. Preparar y guardar en la Base de Datos ---
    datos_para_db = []
    # Formatear la lista de resultados para que coincida con la tabla
    for r in lista_resultados:
        fila = {
            'origen_Departamento': r['origen'].get('Departamento'), 'origen_CUE': r['origen'].get('CUE'),
            'origen_Numero_Escuela': r['origen'].get('Numero_Escuela'), 'origen_Numero_Anexo': r['origen'].get('Numero_Anexo'),
            'origen_Nombre_Escuela': r['origen'].get('Nombre_Escuela'), 'origen_Division': r['origen'].get('Division'),
            'origen_Turno': r['origen'].get('Turno'), 'destino_Departamento': r['destino'].get('Departamento'),
            'destino_CUE': r['destino'].get('CUE'), 'destino_Numero_Escuela': r['destino'].get('Numero_Escuela'),
            'destino_Numero_Anexo': r['destino'].get('Numero_Anexo'), 
            'destino_Nombre_Escuela': r['destino'].get('Nombre_Escuela'),
            'destino_Division': r['destino'].get('Division'), 'destino_Turno': r['destino'].get('Turno'),
            'Distancia_KM': round(r.get('Distancia_KM', 0), 2), 'Observaciones': r.get('Observaciones')
        }
        datos_para_db.append(fila)
    
    # Convertir a DataFrame de Pandas
    df_resultados = pd.DataFrame(datos_para_db)
    
    # Guardar/Reemplazar la tabla de resultados en la BD
    conn = get_db_connection()
    try:
        df_resultados.to_sql('resultados_asignacion', conn, if_exists='replace', index=False)
        conn.commit()
        print(f"INFO: Se han recalculado y guardado {len(df_resultados)} asignaciones.")
    except Exception as e:
        print(f"ERROR: No se pudieron guardar las asignaciones en la BD. {e}")
    finally:
        conn.close()

# --- Endpoints de la API ---

@app.route("/api/asignaciones", methods=['GET'])
def get_asignaciones():
    """
    ENDPOINT PRINCIPAL (GET): /api/asignaciones
    Devuelve al frontend la lista de asignaciones, aplicando filtros y paginación.
    """
    # --- 1. Obtener parámetros de la URL ---
    page = request.args.get('page', 1, type=int)
    limit = request.args.get('limit', 15, type=int)
    offset = (page - 1) * limit
    
    # Filtros
    filtro_depto = request.args.get('departamento')
    filtro_turno = request.args.get('turno')
    filtro_estado = request.args.get('estado_asignacion')
    filtro_nombre = request.args.get('nombre_escuela') # Filtro de búsqueda

    conn = get_db_connection()
    
    # --- 2. Construir la consulta SQL dinámicamente ---
    base_query = "FROM resultados_asignacion"
    where_clause = " WHERE 1=1" # Truco para poder añadir "AND" fácilmente
    params = [] # Lista de parámetros para evitar inyección SQL

    # Añadir filtros a la consulta
    if filtro_depto and filtro_depto != 'todos':
        where_clause += " AND origen_Departamento = ?"
        params.append(filtro_depto)
        
    if filtro_turno and filtro_turno != 'todos':
        where_clause += " AND origen_Turno = ?"
        params.append(filtro_turno)

    if filtro_estado and filtro_estado != 'todos':
        if filtro_estado == '0-5km':
            where_clause += " AND Observaciones = 'Asignado (0-5 km)'"
        elif filtro_estado == '5-10km':
            where_clause += " AND Observaciones = 'Asignado (5-10 km)'"
        elif filtro_estado == '10-30km':
            where_clause += " AND Observaciones = 'Asignado (10-30 km)'"
        elif filtro_estado == 'no-asignadas':
            where_clause += " AND Observaciones LIKE 'No Asignada%'"

    if filtro_nombre:
        # Busca el texto en la columna de origen O en la de destino
        where_clause += " AND (origen_Nombre_Escuela LIKE ? OR destino_Nombre_Escuela LIKE ?)"
        params.append(f"%{filtro_nombre}%") # El % permite búsquedas parciales (contiene)
        params.append(f"%{filtro_nombre}%")

    try:
        # --- 3. Ejecutar Consultas ---
        
        # Consulta para contar el TOTAL de items (para la paginación)
        total_items_query = f"SELECT COUNT(*) {base_query}{where_clause}"
        total_items = conn.execute(total_items_query, params).fetchone()[0]
        
        # Consulta para obtener los datos de la PÁGINA ACTUAL
        data_query = f"SELECT * {base_query}{where_clause} LIMIT ? OFFSET ?"
        query_results = conn.execute(data_query, (*params, limit, offset)).fetchall()
        
        # Consultas para llenar los menús de filtros (selects)
        departamentos_rows = conn.execute("SELECT DISTINCT origen_Departamento FROM resultados_asignacion WHERE origen_Departamento IS NOT NULL ORDER BY origen_Departamento").fetchall()
        turnos_rows = conn.execute("SELECT DISTINCT origen_Turno FROM resultados_asignacion WHERE origen_Turno IS NOT NULL ORDER BY origen_Turno").fetchall()
        
        all_departamentos = [row[0] for row in departamentos_rows]
        all_turnos = [row[0] for row in turnos_rows]

        # --- 4. Devolver JSON al frontend ---
        return jsonify({
            'totalItems': total_items,
            'asignaciones': [dict(row) for row in query_results], # Convierte los resultados a dict
            'totalPages': math.ceil(total_items / limit),
            'currentPage': page,
            'allDepartamentos': all_departamentos,
            'allTurnos': all_turnos
        })
    except Exception as e:
        print(f"ERROR al obtener asignaciones: {e}")
        return jsonify({"error": "No se pudieron obtener las asignaciones."}), 500
    finally:
        conn.close()

@app.route('/api/cargar-planillas', methods=['POST'])
def cargar_planillas():
    """
    ENDPOINT (POST): /api/cargar-planillas
    Recibe uno o más archivos Excel, los envía al 'procesador_excel' para
    limpiarlos y unificarlos, y luego recalcula todas las asignaciones.
    """
    if 'planillas' not in request.files:
        return jsonify({"error": "No se encontraron archivos."}), 400
    
    lista_archivos = request.files.getlist('planillas')
    if not lista_archivos:
        return jsonify({"error": "La lista de archivos está vacía."}), 400
    
    # Carga los datos existentes (si hay) para mergearlos
    conn = get_db_connection()
    try:
        df_existente = pd.read_sql_query("SELECT * FROM escuelas_data", conn)
    except pd.io.sql.DatabaseError:
        df_existente = pd.DataFrame()
        print("INFO: Tabla 'escuelas_data' no encontrada. Se creará una nueva.")
    finally:
        conn.close()

    # Llama al módulo externo para procesar los archivos
    df_actualizado = procesar_archivos_y_actualizar(lista_archivos, df_existente)

    # Si el procesamiento fue exitoso, guarda en la BD y recalcula
    if df_actualizado is not None and not df_actualizado.empty:
        conn = get_db_connection()
        try:
            # Reemplaza la tabla de datos de escuelas con la nueva info
            df_actualizado.to_sql('escuelas_data', conn, if_exists='replace', index=False)
            conn.commit()
            
            # Llama a la función "cerebro" para recalcular todo
            recalcular_y_guardar_asignaciones()
            
            return jsonify({"mensaje": "Éxito: Datos procesados y asignaciones recalculadas."}), 200
        except Exception as e:
            return jsonify({"error": f"Error al guardar o recalcular: {e}"}), 500
        finally:
            if conn: conn.close()
    else:
        print("ADVERTENCIA: El procesamiento de archivos no generó datos válidos.")
        return jsonify({"error": "El procesamiento de los archivos falló o no contenían datos válidos."}), 400

@app.route('/api/descargar-excel', methods=['GET'])
def descargar_excel():
    """
    ENDPOINT (GET): /api/descargar-excel
    Genera y devuelve un archivo Excel con las asignaciones, aplicando filtros
    y estilos profesionales (colores, bordes, auto-ancho).
    """
    # --- 1. Obtener filtros (similar a get_asignaciones) ---
    departamento = request.args.get('departamento')
    turno = request.args.get('turno')
    estado_asignacion = request.args.get('estado_asignacion')

    conn = get_db_connection()
    try:
        # --- 2. Consultar la BD con los filtros ---
        query = "SELECT * FROM resultados_asignacion WHERE 1=1"
        params = []

        if departamento and departamento != 'todos':
            query += " AND origen_Departamento = ?"
            params.append(departamento)
        if turno and turno != 'todos':
            query += " AND origen_Turno = ?"
            params.append(turno)
        
        if estado_asignacion and estado_asignacion != 'todos':
            if estado_asignacion == '0-5km':
                query += " AND Observaciones = 'Asignado (0-5 km)'"
            elif estado_asignacion == '5-10km':
                query += " AND Observaciones = 'Asignado (5-10 km)'"
            elif estado_asignacion == '10-30km':
                query += " AND Observaciones = 'Asignado (10-30 km)'"
            elif estado_asignacion == 'no-asignadas':
                query += " AND Observaciones LIKE 'No Asignada%'"
            
        df = pd.read_sql_query(query, conn, params=tuple(params))
        
        if df.empty:
            return jsonify({"error": "No hay datos para exportar con los filtros seleccionados."}), 404

        # --- 3. Renombrar Columnas para un reporte legible ---
        df.rename(columns={
            'origen_Departamento': 'Origen - Departamento', 'origen_CUE': 'Origen - CUE',
            'origen_Numero_Escuela': 'Origen - N° Escuela', 'origen_Numero_Anexo': 'Origen - Anexo',
            'origen_Nombre_Escuela': 'Origen - Nombre', 'origen_Division': 'Origen - División', 'origen_Turno': 'Origen - Turno',
            'destino_Departamento': 'Destino - Departamento', 'destino_CUE': 'Destino - CUE',
            'destino_Numero_Escuela': 'Destino - N° Escuela', 'destino_Numero_Anexo': 'Destino - Anexo',
            'destino_Nombre_Escuela': 'Destino - Nombre',
            'destino_Division': 'Destino - División', 'destino_Turno': 'Destino - Turno',
            'Distancia_KM': 'Distancia (KM)', 'Observaciones': 'Observaciones'
        }, inplace=True)

        # Definir el orden final de las columnas
        columnas_excel = [
            'Origen - Departamento', 'Origen - CUE', 'Origen - N° Escuela', 'Origen - Anexo',
            'Origen - Nombre', 'Origen - División', 'Origen - Turno',
            'Destino - Departamento', 'Destino - CUE', 'Destino - N° Escuela', 'Destino - Anexo',
            'Destino - Nombre', 'Destino - División', 'Destino - Turno',
            'Distancia (KM)', 'Observaciones'
        ]
        
        columnas_existentes = [col for col in columnas_excel if col in df.columns]
        df = df[columnas_existentes] # Reordenar el DataFrame

        # --- 4. Crear el archivo Excel en Memoria ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Asignaciones')
            
            # --- INICIO DE CÓDIGO DE ESTILO (OPENPYXL) ---
            
            # Obtener la hoja de trabajo (worksheet)
            worksheet = writer.sheets['Asignaciones']

            # --- 4.1. Definir Estilos ---
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            banded_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            thin_border_side = Side(border_style="thin", color="BFBFBF")
            cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")

            # --- 4.2. Aplicar Estilo de Cabecera y Auto-ancho ---
            max_lengths = {} # Diccionario para guardar el ancho máximo de cada columna

            for row_idx, row in enumerate(worksheet.iter_rows(), 1):
                # Es la cabecera (fila 1)
                if row_idx == 1:
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_align
                        cell.border = cell_border
                        max_lengths[cell.column] = len(str(cell.value))
                    continue 
                
                # --- 4.3. Aplicar Estilo de Filas (Bordes, Bandas, Alineación) ---
                is_even_row = row_idx % 2 == 0 
                
                for cell in row:
                    cell.border = cell_border
                    
                    if is_even_row: # Aplicar banda gris a filas pares
                        cell.fill = banded_fill

                    # Lógica de Alineación (Nombres a la izq, números a la der, resto centrado)
                    col_letter = get_column_letter(cell.column)
                    col_name = worksheet[f"{col_letter}1"].value 
                    
                    if col_name in ['Origen - Nombre', 'Destino - Nombre', 'Observaciones']:
                        cell.alignment = align_left
                    elif col_name == 'Distancia (KM)':
                        cell.alignment = align_right
                    else:
                        cell.alignment = align_center
                    
                    # Calcular ancho máximo para la columna
                    cell_len = len(str(cell.value))
                    if cell.column not in max_lengths or cell_len > max_lengths[cell.column]:
                        max_lengths[cell.column] = cell_len

            # --- 4.4. Ajustar Ancho de Columnas ---
            for col_idx, max_len in max_lengths.items():
                col_letter = get_column_letter(col_idx)
                adjusted_width = min(max_len + 2, 50) # Ancho = max(contenido) + 2 de padding
                worksheet.column_dimensions[col_letter].width = adjusted_width
            
            # --- FIN DE CÓDIGO DE ESTILO ---

        output.seek(0) # Regresa al inicio del archivo en memoria
        
        # --- 5. Enviar Archivo al Usuario ---
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='reporte_asignaciones.xlsx'
        )
    except Exception as e:
        print(f"Error al generar el Excel: {e}")
        return jsonify({"error": "No se pudo generar el archivo Excel."}), 500
    finally:
        if conn: conn.close()

# --- Endpoints de Gestión de Usuarios ---

@app.route("/api/usuarios/rol", methods=['GET'])
def get_user_role():
    """
    ENDPOINT (GET): /api/usuarios/rol
    Se llama al iniciar sesión. Recibe un email (de Auth0) y devuelve el rol
    del usuario guardado en la BD.
    Si el usuario es nuevo, lo crea con rol 'profesor-pendiente'.
    """
    user_email = request.args.get('email')
    if not user_email: return jsonify({"error": "Email no proporcionado"}), 400
    conn = get_db_connection()
    try:
        user = conn.execute('SELECT rol FROM usuarios WHERE email = ?', (user_email,)).fetchone()
        if user:
            # Si el usuario existe, devuelve su rol
            return jsonify({"rol": user['rol']})
        else:
            # Si es la primera vez que inicia sesión, lo crea como 'pendiente'
            conn.execute("INSERT INTO usuarios (email, password_hash, rol) VALUES (?, ?, ?)", (user_email, 'auth0', 'profesor-pendiente'))
            conn.commit()
            return jsonify({"rol": "profesor-pendiente"})
    finally:
        conn.close()

@app.route("/api/usuarios", methods=['GET'])
def get_all_users():
    """
    ENDPOINT (GET): /api/usuarios
    (Solo para Admins) Devuelve la lista de todos los usuarios para el panel de admin.
    """
    conn = get_db_connection()
    try:
        users = conn.execute("SELECT id, email, rol FROM usuarios").fetchall()
        return jsonify([dict(user) for user in users])
    finally:
        conn.close()

@app.route("/api/usuarios/cambiar-rol", methods=['POST'])
def change_user_role():
    """
    ENDPOINT (POST): /api/usuarios/cambiar-rol
    (Solo para Admins) Recibe un ID de usuario y un nuevo rol, y lo actualiza en la BD.
    """
    data = request.json
    user_id = data.get('id')
    new_role = data.get('rol')
    conn = get_db_connection()
    try:
        conn.execute("UPDATE usuarios SET rol = ? WHERE id = ?", (new_role, user_id))
        conn.commit()
        return jsonify({"mensaje": f"Rol del usuario {user_id} actualizado a {new_role}."})
    finally:
        conn.close()

# --- Punto de Entrada Principal ---
if __name__ == "__main__":
    """
    Esto solo se ejecuta cuando corres el archivo directamente (ej: python asignador.py)
    Inicia el servidor de desarrollo de Flask.
    """
    app.run(debug=True, port=5000)