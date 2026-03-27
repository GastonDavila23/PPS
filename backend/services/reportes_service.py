import pandas as pd
import io
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

def generar_excel_estilizado(df):
    output = io.BytesIO()
    
    # 1. Limpieza de columnas de sistema
    cols_a_quitar = ['id_resultado', 'id_carga', 'slot_id']
    df_export = df.drop(columns=[c for c in cols_a_quitar if c in df.columns]).copy()

    # Colores para los rangos de asignación
    fill_verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fill_azul = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    fill_amarillo = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_naranja = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Mapa de anchos
    mapa_anchos = {
        'origen_Departamento': 12, 'destino_Departamento': 12,
        'origen_CUE': 12, 'destino_CUE': 12,
        'origen_Numero_Escuela': 10, 'destino_Numero_Escuela': 10,
        'origen_Numero_Anexo': 10, 'destino_Numero_Anexo': 10,
        'origen_Turno': 10, 'destino_Turno': 10,
        'origen_Nombre_Escuela': 28, 'destino_Nombre_Escuela': 28,
        'Observaciones': 16,
        'origen_Division': 6, 'destino_Division': 6,
        'Distancia_KM': 8
    }

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Asignaciones_DGE')
        worksheet = writer.sheets['Asignaciones_DGE']
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        border_fino = Border(left=Side(style='thin'), right=Side(style='thin'), 
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Aplicar Cabecera y Anchos
        for i, column in enumerate(df_export.columns):
            cell = worksheet.cell(row=1, column=i+1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border_fino

            letter = cell.column_letter
            worksheet.column_dimensions[letter].width = mapa_anchos.get(column, 15)

        obs_col_idx = 1
        for i, col in enumerate(df_export.columns):
            if col == "Observaciones":
                obs_col_idx = i + 1
                break

        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            obs_value = str(row[obs_col_idx-1].value) if row[obs_col_idx-1].value else ""
            
            fill_actual = None
            if "< 1 km" in obs_value: fill_actual = fill_verde
            elif "1-5 km" in obs_value: fill_actual = fill_azul
            elif "5-10 km" in obs_value: fill_actual = fill_amarillo
            elif "10-30 km" in obs_value: fill_actual = fill_naranja

            for cell in row:
                if fill_actual:
                    cell.fill = fill_actual
                
                cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                cell.border = border_fino
                cell.font = Font(size=10)

        worksheet.freeze_panes = 'A2'

    output.seek(0)
    return output